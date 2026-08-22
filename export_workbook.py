"""Excel export — the analysis as real cells, not a picture of a table.

WHY THIS IS NOT JUST THE DECK IN ANOTHER FORMAT. A deck is for showing; a
workbook is for continuing. An analyst who receives this should be able
to sort, filter, pivot and re-model — so every figure lands in its own
cell as a REAL NUMBER with a number format, never as a pre-formatted
string like "27.62%". A string looks identical on screen and is useless
in a formula, which is the single most common way an "Excel export" turns
out to be a spreadsheet-shaped PDF.

MISSING DATA IS A BLANK CELL PLUS A STATUS COLUMN, and that combination
is deliberate. Writing 0 for an unavailable metric would corrupt every
SUM and AVERAGE downstream. A blank alone behaves correctly in formulas —
Excel skips it rather than treating it as zero — but tells the reader
nothing. So the number cell stays empty and an adjacent Status column
says "not reported": correct arithmetic and an honest reader experience,
which neither choice delivers alone.

THRESHOLDS TRAVEL WITH THE VALUES. Each scorecard row carries the
threshold it was judged against, so the workbook is self-contained: a
recipient can see WHY something passed without opening the app, and can
change a threshold and re-derive the verdict themselves.
"""
import datetime
import io
import logging
from dataclasses import dataclass, field
from typing import List, Optional, Sequence, Tuple

from logging_setup import get_logger, log_event, log_exception

logger = get_logger("export_workbook")

_NOT_REPORTED = "not reported"

# Column widths, in Excel's character units. Set explicitly because the
# default cuts off every metric label, and a workbook that needs manual
# column-dragging before it is readable looks unfinished.
_WIDTHS = {"label": 34, "value": 14, "unit": 10, "status": 16, "detail": 52}


@dataclass(frozen=True)
class Row:
    """One metric. `value` of None means unavailable — the cell is left
    blank so formulas skip it, and status says why.

    `percent=True` means `value` is percent-valued (27.62 for 27.62%),
    which is what every function in this app returns. Excel disagrees:
    its percent formats MULTIPLY the stored number by 100 to display it,
    so storing 27.62 under "0.00%" renders as 2762.00%. The writer stores
    value/100 for percent rows so the cell both DISPLAYS 27.62% and
    behaves like a percentage in formulas. Callers keep using the app's
    convention and this is handled in exactly one place.
    """
    label: str
    value: Optional[float] = None
    unit: str = ""
    detail: str = ""
    percent: bool = False
    number_format: str = "#,##0.00"

    @property
    def status(self) -> str:
        return "ok" if self.value is not None else _NOT_REPORTED

    @property
    def cell_value(self) -> Optional[float]:
        """The number as Excel must store it to display correctly."""
        if self.value is None:
            return None
        return float(self.value) / 100.0 if self.percent else float(self.value)

    @property
    def cell_format(self) -> str:
        if self.percent:
            # Excel renders this as e.g. 27.62%
            return "0.00%"
        return self.number_format


@dataclass(frozen=True)
class Sheet:
    title: str
    rows: Tuple[Row, ...] = ()
    note: str = ""


@dataclass(frozen=True)
class WorkbookData:
    ticker: str
    company_name: str = ""
    sector: str = ""
    as_of: Optional[datetime.date] = None
    sheets: Tuple[Sheet, ...] = ()
    summary_lines: Tuple[str, ...] = ()


def is_available() -> Tuple[bool, Optional[str]]:
    try:
        import openpyxl  # noqa: F401
        return True, None
    except Exception:
        return False, (
            "Excel export needs the openpyxl package. Run "
            "`pip install -r requirements.txt` and restart Streamlit."
        )


def _accent():
    from branding import brand

    accent = (brand().accent_color or "#1F6FEB").lstrip("#")
    if len(accent) == 3:
        accent = "".join(c * 2 for c in accent)
    return accent


def build_workbook(data: WorkbookData) -> Tuple[Optional[bytes], Optional[str]]:
    """The finished .xlsx as bytes. Returns (data, error); one is None.
    Never raises."""
    ok, reason = is_available()
    if not ok:
        return None, reason

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from branding import brand

        accent = _accent()
        header_fill = PatternFill("solid", fgColor=accent)
        header_font = Font(bold=True, color="FFFFFF")
        as_of = data.as_of or datetime.date.today()

        book = Workbook()
        overview = book.active
        overview.title = "Overview"

        overview["A1"] = f"{brand().name} — {data.ticker}"
        overview["A1"].font = Font(bold=True, size=16)
        overview["A2"] = data.company_name or ""
        overview["A3"] = data.sector or ""
        overview["A4"] = "Prepared"
        overview["B4"] = as_of
        overview["B4"].number_format = "yyyy-mm-dd"

        line = 6
        for text in data.summary_lines:
            overview.cell(row=line, column=1, value=text)
            line += 1

        line += 1
        overview.cell(row=line, column=1, value=(
            "Not investment advice. Generated from public market data for research purposes. "
            "A blank value cell means the figure was not reported — it is not a zero, and is "
            "skipped by SUM and AVERAGE rather than dragging them down."
        )).alignment = Alignment(wrap_text=True, vertical="top")
        overview.column_dimensions["A"].width = 96

        for sheet in data.sheets:
            # Excel caps sheet names at 31 characters and rejects several
            # punctuation marks outright; a long section title would
            # otherwise fail the whole save.
            ws = book.create_sheet(_safe_sheet_name(sheet.title))
            headers = ("Metric", "Value", "Unit", "Status", "Detail")
            for column, title in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=column, value=title)
                cell.fill = header_fill
                cell.font = header_font

            for index, row in enumerate(sheet.rows, start=2):
                ws.cell(row=index, column=1, value=row.label)
                if row.value is not None:
                    # A real number with a format — NOT a formatted string.
                    value_cell = ws.cell(row=index, column=2, value=row.cell_value)
                    value_cell.number_format = row.cell_format
                ws.cell(row=index, column=3, value=row.unit)
                ws.cell(row=index, column=4, value=row.status)
                ws.cell(row=index, column=5, value=row.detail)

            for column, key in enumerate(("label", "value", "unit", "status", "detail"), start=1):
                ws.column_dimensions[get_column_letter(column)].width = _WIDTHS[key]
            ws.freeze_panes = "A2"

            if sheet.note:
                note_row = len(sheet.rows) + 3
                ws.cell(row=note_row, column=1, value=sheet.note).alignment = Alignment(
                    wrap_text=True, vertical="top")

        buffer = io.BytesIO()
        book.save(buffer)
        payload = buffer.getvalue()
        log_event(logger, logging.INFO, "export_workbook.built",
                  ticker=data.ticker, sheets=len(data.sheets), bytes=len(payload))
        return payload, None
    except Exception as e:
        log_exception(logger, "export_workbook.build_failed", section="export_workbook")
        return None, f"Couldn't build the workbook ({type(e).__name__})."


def _safe_sheet_name(title: str) -> str:
    """Excel rejects : \\ / ? * [ ] and caps names at 31 characters."""
    cleaned = "".join(c for c in (title or "Sheet") if c not in ':\\/?*[]')
    return (cleaned.strip() or "Sheet")[:31]


def filename_for(ticker: str, as_of: Optional[datetime.date] = None) -> str:
    as_of = as_of or datetime.date.today()
    from branding import brand
    slug = brand().name.lower().replace(" ", "-")
    return f"{slug}-{ticker.lower()}-{as_of:%Y%m%d}.xlsx"
